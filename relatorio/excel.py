from openpyxl import Workbook
from django.http import HttpResponse
from .models import Ocorrencia, DiscadorOcorrencia, Sistema, Carteira

def exporta_ocorrencia(request):
    workbook = Workbook()

    dados_ocorrencia = Ocorrencia.objects.all()

    planilha_ocorrencia = workbook.create_sheet(title='Ocorrências')

    planilha_ocorrencia.append(['pk_interna', 'num_ocorrencia', 'desc_ocorrencia'])

    for indice, dado in enumerate(dados_ocorrencia, start=2):
        planilha_ocorrencia.cell(row=indice, column=1, value=dado.pk_interna)
        planilha_ocorrencia.cell(row=indice, column=2, value=dado.num_ocorrencia)
        planilha_ocorrencia.cell(row=indice, column=3, value=dado.desc_ocorrencia)

    workbook.remove(workbook['Sheet'])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=exportacao.xlsx'

    workbook.save(response)

    return response


def exporta_sistema(request):
    workbook = Workbook()

    dados_sistema = Sistema.objects.all()

    planilha_sistema = workbook.create_sheet(title='Sistemas')

    planilha_sistema.append(['codigo', 'nome_sistema', ])

    for indice, dado in enumerate(dados_sistema, start=2):
        planilha_sistema.cell(row=indice, column=1, value=dado.codigo)
        planilha_sistema.cell(row=indice, column=2, value=dado.nome_sistema)

    workbook.remove(workbook['Sheet'])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=exportacao.xlsx'

    workbook.save(response)

    return response



def exporta_carteira(request):
    workbook = Workbook()

    dados_carteira = Carteira.objects.all()

    planilha_carteira = workbook.create_sheet(title='Carteiras')

    planilha_carteira.append(['cod_carteira', 'nome_carteira', ])

    for indice, dado in enumerate(dados_carteira, start=2):
        planilha_carteira.cell(row=indice, column=1, value=dado.cod_carteira)
        planilha_carteira.cell(row=indice, column=2, value=dado.nome_carteira)

    workbook.remove(workbook['Sheet'])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=exportacao.xlsx'

    workbook.save(response)

    return response


def exporta_classificacao(request):
    workbook = Workbook()

    dados_classificacao = DiscadorOcorrencia.objects.all()

    planilha_classificacao = workbook.create_sheet(title='Classificação')

    planilha_classificacao.append(['Sistema', 'Carteira', 'Ocorrencia', 'alo', 'cpc', 'promessa', ])

    for indice, dado in enumerate(dados_classificacao, start=2):
        planilha_classificacao.cell(row=indice, column=1, value=dado.sist.nome_sistema)
        planilha_classificacao.cell(row=indice, column=2, value=dado.carteira.nome_carteira)
        planilha_classificacao.cell(row=indice, column=3, value=dado.ocorrencia.desc_ocorrencia)
        planilha_classificacao.cell(row=indice, column=4, value=dado.alo)
        planilha_classificacao.cell(row=indice, column=5, value=dado.cpc)
        planilha_classificacao.cell(row=indice, column=6, value=dado.promessa)

    workbook.remove(workbook['Sheet'])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=exportacao.xlsx'

    workbook.save(response)

    return response

